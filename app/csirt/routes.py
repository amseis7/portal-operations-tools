from sqlalchemy import func
from flask import render_template, request, flash, redirect, url_for, send_file, stream_with_context
from flask_login import login_required, current_user
from app.csirt import bp
from app.models import Alerta, Ioc, VtIoc, VtTicket
from app.extensions import db
import csv
import re
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from werkzeug.wrappers import Response

import pandas as pd
from io import StringIO, BytesIO
from datetime import datetime, timedelta

from app.csirt.logic import ejecutar_proceso_csirt, actualizar_iocs_faltantes, generador_actualizacion_masiva, obtener_mapa_recurrencia
from app.utils import admin_required, check_access


@bp.route('/')
@login_required
@check_access('csirt')
def index():
    tickets = db.session.query(
        Alerta.ticket,
        Alerta.responsable,
        func.max(Alerta.fecha_realizacion).label('ultima_fecha'),
        func.count(Alerta.id).label('total_alertas')
    ).group_by(Alerta.ticket, Alerta.responsable).order_by(func.max(Alerta.fecha_realizacion).desc()).all()
    
    return render_template('csirt/index.html', tickets=tickets, titulo_navbar="Gestión CSIRT")

@bp.route('/gestion/<ticket_id>')
@login_required
@check_access('csirt')
def ver_gestion(ticket_id):
    # Buscamos todas las alertas que tengan ese ticket
    alertas = Alerta.query.filter(
        Alerta.ticket == ticket_id,
        Alerta.tipo_alerta != 'AVC' 
    ).order_by(Alerta.fecha_realizacion.desc()).all()
    return render_template('csirt/detalle_gestion.html', alertas=alertas, ticket=ticket_id, titulo_navbar="Gestión CSIRT")

@bp.route('/iocs/<ticket_id>')
@login_required
@check_access('csirt')
def ver_iocs(ticket_id):
    # 1. Query Base
    query = db.session.query(Ioc).join(Alerta).filter(Alerta.ticket == ticket_id)
    
    # 2. Filtro Visual (Si viene en la URL)
    tipo_filtro = request.args.get('tipo')
    
    if tipo_filtro:
        if tipo_filtro == 'hash':
            query = query.filter(Ioc.tipo.in_(['hash', 'md5', 'sha1', 'sha256']))
        elif tipo_filtro == 'url':
            query = query.filter(Ioc.tipo.in_(['url', 'dominio'])) # Agrupamos URL y Dominio
        else:
            query = query.filter(Ioc.tipo == tipo_filtro)
            
    iocs = query.all()

    valores_en_pantalla = [ioc.valor for ioc in iocs]

    mapa_recurrencia = obtener_mapa_recurrencia

    if valores_en_pantalla:
        stats = db.session.query(Ioc.valor, func.count(Ioc.id))\
            .filter(Ioc.valor.in_(valores_en_pantalla))\
            .group_by(Ioc.valor).all()
        
        mapa_recurrencia = {item[0]: item[1] for item in stats}
    
    return render_template(
        'csirt/detalle_iocs.html', 
        iocs=iocs, 
        ticket=ticket_id, 
        titulo=f"Ticket {ticket_id}",
        contexto='ticket', 
        id_contexto=ticket_id,
        filtro_activo=tipo_filtro, # <--- Pasamos esto para saber si mostrar botón "Borrar filtro"
        mapa_recurrencia=mapa_recurrencia
    )

@bp.route('/procesar', methods=['POST'])
@login_required
@check_access('csirt')
def procesar():
    ticket = request.form.get('ticket')

    # Detectar si el checkbox fue marcado (si no está marcado, devuelve None)
    modo_prueba = request.form.get('modo_prueba') == 'on'

    if not ticket or not re.match(r'^RF-\d+$', ticket):
        flash('Error: El ticket debe tener el formato RF-xxxxxx (Ej: RF-123456)', 'danger')
        return redirect(url_for('csirt.index'))
    
    # Llamamos a la lógica maestra
    responsable_real = current_user.nombre_completo or current_user.username
    resultado = ejecutar_proceso_csirt(ticket, responsable_real, simulacion=modo_prueba)
    
    if resultado['status'] == 'exito':
        flash(f"¡Proceso finalizado! {resultado['msg']}", 'success' if not modo_prueba else 'warning')
    else:
        flash(f"Proceso finalizado. {resultado['msg']}", 'info')
    
    return redirect(url_for('csirt.index'))

@bp.route('/actualizar_iocs/<ticket_id>', methods=['POST'])
@login_required
@check_access('csirt')
def actualizar_iocs(ticket_id):
    try:
        cant_alertas, cant_iocs = actualizar_iocs_faltantes(ticket_id)
        
        if cant_alertas > 0:
            flash(f'Se actualizaron {cant_alertas} alertas con un total de {cant_iocs} IoCs nuevos.', 'success')
        else:
            flash('No se encontraron IoCs nuevos o las alertas ya estaban completas.', 'info')
            
    except Exception as e:
        flash(f'Ocurrió un error al actualizar: {str(e)}', 'danger')
        
    # Volvemos a la misma página de gestión
    return redirect(url_for('csirt.ver_gestion', ticket_id=ticket_id))

# En app/csirt/routes.py

@bp.route('/admin/exportar_todo_csv')
@login_required
@check_access('csirt')
def exportar_todo_csv():
    # 1. Seguridad: Solo admin
    if not current_user.is_admin:
        flash('Acceso denegado.', 'danger')
        return redirect(url_for('csirt.index'))

    # 2. Consultar TODAS las alertas
    alertas = Alerta.query.order_by(Alerta.fecha_realizacion.desc()).all()

    # 3. Crear CSV en memoria
    si = StringIO()
    # Usamos punto y coma (;) para compatibilidad con tu importador
    cw = csv.writer(si, delimiter=';') 
    
    # Encabezados (Deben coincidir con lo que espera tu importador)
    cw.writerow(['Nombre Alerta', 'Tipo', 'Ticket', 'Responsable', 'Fecha'])

    # 4. Escribir datos
    for alerta in alertas:
        cw.writerow([
            alerta.nombre_alerta,
            alerta.tipo_alerta,
            alerta.ticket,
            alerta.responsable,
            alerta.fecha_realizacion.strftime('%d-%m-%Y') # Formato dd-mm-yyyy
        ])

    # 5. Descargar
    output = si.getvalue()
    fecha_hoy = datetime.now().strftime('%Y%m%d')
    
    return Response(
        output,
        mimetype="text/csv",
        headers={"Content-disposition": f"attachment; filename=Respaldo_Total_CSIRT_{fecha_hoy}.csv"}
    )

@bp.route('/importar_historico', methods=['GET', 'POST'])
@login_required
@admin_required
def importar_historico():
    # --- Verificacion de seguridad --- #
    #if not current_user.is_admin:
    #    flash('Acceso denegado. Se requieren permisos de administracion.', 'danger')
    #    return redirect(url_for('csirt.index'))
    # --------------------------------- #
    if request.method == 'POST':
        file = request.files.get('archivo_csv')
        if not file:
            flash('sube un archivo CSV', 'warning')
            return redirect(request.url)
        
        try:
            # 1. Solución al error de codificación (UTF-8 vs Latin-1)
            file_bytes = file.stream.read()
            try:
                # Intentamos UTF-8 primero (estándar web)
                content = file_bytes.decode('utf-8')
            except UnicodeDecodeError:
                # Si falla, usamos Latin-1 (estándar Excel español)
                content = file_bytes.decode('latin-1')

            stream = StringIO(content)
            
            # 2. Solución al separador: Usamos sep=';'
            df = pd.read_csv(stream, sep=';')
            
            # Limpieza de nombres de columnas (quita espacios extra si los hay)
            df.columns = df.columns.str.strip()
            
            contador = 0
            for index, row in df.iterrows():
                # Mapeo de columnas basado en tu archivo real
                nombre_alerta = row.get('Nombre Alerta')
                
                # Evitamos filas vacías
                if pd.isna(nombre_alerta):
                    continue

                existe = Alerta.query.filter_by(nombre_alerta=nombre_alerta).first()
                
                if not existe:
                    # Procesamiento de Fecha (Tu archivo tiene formato dd-mm-yyyy)
                    fecha_str = str(row.get('Fecha', ''))
                    try:
                        # Intentamos varios formatos por seguridad
                        if '-' in fecha_str:
                            fecha_obj = datetime.strptime(fecha_str, '%d-%m-%Y')
                        else:
                            fecha_obj = datetime.strptime(fecha_str, '%d/%m/%Y')
                    except ValueError:
                        fecha_obj = datetime.now()

                    nueva_alerta = Alerta(
                        nombre_alerta=nombre_alerta,
                        ticket=row.get('Ticket', 'S/T'),
                        tipo_alerta=row.get('Tipo', 'N/A'),
                        responsable=row.get('Responsable', 'Sistema'),
                        fecha_realizacion=fecha_obj
                    )
                    db.session.add(nueva_alerta)
                    contador += 1
            
            db.session.commit()
            flash(f'Éxito: Se importaron {contador} alertas nuevas.', 'success')
            return redirect(url_for('csirt.index'))

        except Exception as e:
            db.session.rollback()
            # Muestra el error detallado para depurar
            flash(f'Error al importar: {str(e)}', 'danger')

    return render_template('csirt/importar.html')

# --- NUEVA RUTA: Descargar CSV Masivo por Ticket ---
@bp.route('/descargar_csv/<ticket_id>')
@login_required
@check_access('csirt')
def descargar_iocs_csv(ticket_id):
    # 1. Buscamos todos los IoCs asociados a alertas de este ticket
    iocs = db.session.query(Ioc).join(Alerta).filter(Alerta.ticket == ticket_id).order_by(Ioc.tipo).all()
    
    if not iocs:
        flash(f'No hay IoCs registrados para el ticket {ticket_id}', 'warning')
        return redirect(url_for('csirt.index'))

    # 2. Crear el archivo CSV en memoria (StringIO)
    si = StringIO()
    cw = csv.writer(si, delimiter=';') # Usamos punto y coma para Excel en español
    
    # Encabezados
    cw.writerow(['Tipo', 'Valor', 'Alerta Origen', 'Ticket'])
    
    # Filas
    for ioc in iocs:
        cw.writerow([ioc.tipo, ioc.valor, ioc.alerta.nombre_alerta, ticket_id])
    
    # 3. Crear la respuesta HTTP para descarga
    output = si.getvalue()
    return Response(
        output,
        mimetype="text/csv",
        headers={"Content-disposition": f"attachment; filename=IoCs_{ticket_id}.csv"}
    )

# --- NUEVA RUTA: Ver IoCs de una ALERTA ESPECÍFICA ---
@bp.route('/iocs_alerta/<int:alerta_id>')
@login_required
@check_access('csirt')
def ver_iocs_alerta(alerta_id):
    alerta = Alerta.query.get_or_404(alerta_id)
    
    # Query Base sobre los IoCs de esta alerta
    # Nota: Usamos alerta.iocs (lista) o construimos query si queremos filtrar
    query = Ioc.query.filter_by(alerta_id=alerta_id)
    
    # Filtro Visual
    tipo_filtro = request.args.get('tipo')
    
    if tipo_filtro:
        if tipo_filtro == 'hash':
            query = query.filter(Ioc.tipo.in_(['hash', 'md5', 'sha1', 'sha256']))
        elif tipo_filtro == 'url':
            query = query.filter(Ioc.tipo.in_(['url', 'dominio']))
        else:
            query = query.filter(Ioc.tipo == tipo_filtro)
            
    iocs = query.all()

    mapa_recurrencia =obtener_mapa_recurrencia(iocs)

    return render_template(
        'csirt/detalle_iocs.html', 
        iocs=iocs, 
        ticket=alerta.ticket,
        titulo=f"Alerta {alerta.nombre_alerta}", 
        volver='gestion', 
        id_volver=alerta.ticket,
        contexto='alerta', 
        id_contexto=alerta.id,
        filtro_activo=tipo_filtro, # <--- Nuevo
        mapa_recurrencia=mapa_recurrencia
    )

@bp.route('/eliminar_ticket/<ticket_id>', methods=['POST'])
@login_required
@admin_required
def eliminar_ticket(ticket_id):
    # 1. Seguridad: Solo admin puede borrar
    try:
        # 2. Buscar todas las alertas de este ticket
        alertas_a_borrar = Alerta.query.filter_by(ticket=ticket_id).all()
        cantidad = len(alertas_a_borrar)
        
        if cantidad == 0:
            flash('No se encontraron registros para este ticket.', 'warning')
            return redirect(url_for('csirt.index'))

        # 3. Borrarlas (Los IoCs se borran en cascada automáticamente)
        for alerta in alertas_a_borrar:
            db.session.delete(alerta)
        
        db.session.commit()
        
        flash(f'Se eliminó el ticket {ticket_id} y sus {cantidad} alertas asociadas (con sus IoCs).', 'success')

    except Exception as e:
        db.session.rollback()
        flash(f'Error al eliminar: {str(e)}', 'danger')

    return redirect(url_for('csirt.index'))

@bp.route('/generar_reporte', methods=['POST'])
@login_required
@check_access('csirt')
def generar_reporte():
    # 1. Obtener fechas
    fecha_inicio_str = request.form.get('fecha_inicio')
    fecha_fin_str = request.form.get('fecha_fin')

    if not fecha_inicio_str or not fecha_fin_str:
        flash('Debes seleccionar ambas fechas.', 'warning')
        return redirect(url_for('csirt.index'))

    try:
        fecha_inicio = datetime.strptime(fecha_inicio_str, '%Y-%m-%d').date()
        fecha_fin = datetime.strptime(fecha_fin_str, '%Y-%m-%d').date()
        fecha_fin_inclusive = datetime.combine(fecha_fin, datetime.max.time())
    except ValueError:
        flash('Formato de fecha inválido.', 'danger')
        return redirect(url_for('csirt.index'))

    # 2. Consultar BD
    alertas = Alerta.query.filter(
        Alerta.fecha_realizacion >= fecha_inicio,
        Alerta.fecha_realizacion <= fecha_fin_inclusive
    ).order_by(Alerta.fecha_realizacion.asc()).all()

    if not alertas:
        flash(f'No se encontraron alertas entre {fecha_inicio_str} y {fecha_fin_str}.', 'info')
        return redirect(url_for('csirt.index'))

    # 3. Preparar Excel
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Reporte Gestión"

    # --- ESTILOS ---
    header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", name="Calibri", size=11)
    center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
    left_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'), 
        top=Side(style='thin'), bottom=Side(style='thin')
    )

    # --- ENCABEZADOS (Ordenados según tu solicitud) ---
    headers = [
        'Nº',
        'Fecha',
        'Detalle Incidente / Reporte de vulnerabilidades',
        'Tipo (Phishing, virus, etc.)',
        'Criticidad (Alta, Media, Baja)',
        'Canal de información de obtención (Twitter, email, etc.)',
        'Plan de acción',
        'Fecha de implementación',
        'Chequeo post remediación'
    ]
    ws.append(headers)

    # Estilar Cabecera y definir anchos
    column_widths = [5, 12, 40, 25, 15, 25, 25, 20, 25] # Anchos para col 1 a 9
    
    for col_num, cell in enumerate(ws[1], 1):
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center_align
        cell.border = thin_border
        # Aplicar anchos
        col_letter = openpyxl.utils.get_column_letter(col_num)
        if col_num <= len(column_widths):
            ws.column_dimensions[col_letter].width = column_widths[col_num-1]

    # --- DICCIONARIO DE TIPOS ---
    mapa_tipos = {
        "8FPH": "Phishing",
        "2CMV": "Malware",
        "8FFR": "Sitios Fraudulentos",
        "4IIA": "Ataques de Fuerza Bruta",
        "4IIV": "Ataques de Fuerza Bruta",
        "ACF": "Campaña Fraudulenta",
        "AIA": "Investigacion de Amenazas",
        "AVC": "Vulnerabilidad Critica"
    }

    # --- LLENADO DE DATOS ---
    contador = 1
    for alerta in alertas:
        
        # A. Lógica de Traducción de Tipo
        # Buscamos si alguna de las claves (Ej: 8FPH) está contenida en el nombre de la alerta
        tipo_reporte = "Otro / Desconocido"
        nombre_upper = alerta.nombre_alerta.upper()
        
        for codigo, descripcion in mapa_tipos.items():
            if codigo in nombre_upper:
                tipo_reporte = descripcion
                break
        
        # Si no encontró coincidencias exactas pero tenemos el tipo corto guardado
        if tipo_reporte == "Otro / Desconocido" and alerta.tipo_alerta:
             # Intento de fallback (Ej: si la BD tiene 'AIA' pero el nombre no)
             tipo_reporte = mapa_tipos.get(alerta.tipo_alerta, alerta.tipo_alerta)

        # C. Construcción de la Fila
        row_data = [
            contador,                                      # Nº
            alerta.fecha_realizacion.strftime('%d/%m/%Y'), # Fecha
            alerta.nombre_alerta,                          # Detalle
            tipo_reporte,                                  # Tipo Traducido
            "Alta",                                        # Criticidad
            "CSIRT",                                       # Canal
            "Bloqueo de IoC",                              # Plan de acción (Vacio para llenar manual)
            alerta.fecha_realizacion.strftime('%d/%m/%Y'), # Fecha implementación (Vacio)
            "OK"                                             # Chequeo post (Vacio)
        ]
        ws.append(row_data)
        
        # Estilar celdas de datos
        current_row = ws.max_row
        for col_idx, cell in enumerate(ws[current_row], 1):
            cell.border = thin_border
            # Alinear a la izquierda el Detalle y el Tipo, el resto centrado
            if col_idx in [3, 4, 7, 9]: 
                cell.alignment = left_align
            else:
                cell.alignment = center_align
        
        contador += 1

    # Guardar y enviar
    excel_file = BytesIO()
    wb.save(excel_file)
    excel_file.seek(0)

    nombre_archivo = f"Reporte_CSIRT_{fecha_inicio_str}_al_{fecha_fin_str}.xlsx"
    return send_file(
        excel_file,
        as_attachment=True,
        download_name=nombre_archivo,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )

@bp.route('/admin/stream_actualizacion')
@login_required
@admin_required
def stream_actualizacion():
    # Esta respuesta mantiene la conexión abierta y envía texto plano
    return Response(stream_with_context(generador_actualizacion_masiva()), mimetype='text/plain')

@bp.route('/buscar', methods=['GET'])
@login_required
def buscar_ioc():
    query_str = request.args.get('q', '').strip()

    if not query_str:
        flash('Por favor ingrese un termino de busqueda.', 'warning')
        return redirect(url_for('csirt.index'))
    
    # 1. Búsqueda en CSIRT (Histórico de Tickets/Alertas)
    resultados_csirt = db.session.query(Ioc)\
        .join(Alerta)\
        .filter(Ioc.valor.contains(query_str))\
        .order_by(Alerta.fecha_realizacion.desc())\
        .all()

    # 2. Búsqueda en Casos VT (Investigaciones)
    resultados_vt = db.session.query(VtIoc)\
        .join(VtTicket)\
        .filter(VtIoc.valor.contains(query_str))\
        .order_by(VtTicket.fecha_creacion.desc())\
        .all()
    
    # Enviamos ambas listas a la plantilla
    return render_template(
        'csirt/resultados_busqueda.html', 
        resultados_csirt=resultados_csirt, 
        resultados_vt=resultados_vt, 
        query=query_str
    )