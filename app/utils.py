from functools import wraps
from flask import flash, redirect, url_for
from flask_login import current_user, login_required
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from io import BytesIO
from datetime import datetime

def admin_required(f):
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if not current_user.is_admin:
            #abort(403)
            flash('Acceso denegado.', 'danger')
            return redirect(url_for('main.dashboard'))
        return f(*args, **kwargs)
    return decorated_function

def proteger_blueprint(bp, nombre_permiso):
    """
    Aplica seguridad a todo un blueprint de forma autentica.
    """
    @bp.before_request
    @login_required
    def verificar_acceso():
        if not current_user.has_tool(nombre_permiso):
            flash(f'No tienes acceso al módulo de {nombre_permiso}.', 'warning')
            return redirect(url_for('main.dashboard'))

def generar_reporte_excel(alertas, fecha_inicio_str, fecha_fin_str):
    """
    Genera un reporte Excel a partir de una lista de alertas.
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Reporte Gestión"

    header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", name="Calibri", size=11)
    center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
    left_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'), 
        top=Side(style='thin'), bottom=Side(style='thin')
    )

    headers = [
        'Nº',
        'Fecha',
        'Detalle Incidente / Reporte de vulnerabilidades',
        'Tipo (Phishing, virus, etc.)',
        'Criticidad (Alta, Media, Baja)',
        'Canal de información de obtención (Twitter, email, etc.)',
        'Plan de acción',
        'Fecha de implementación',
        'Chequeo post remediación'
    ]
    ws.append(headers)

    column_widths = [5, 12, 40, 25, 15, 25, 25, 20, 25]
    
    for col_num, cell in enumerate(ws[1], 1):
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center_align
        cell.border = thin_border
        col_letter = openpyxl.utils.get_column_letter(col_num)
        if col_num <= len(column_widths):
            ws.column_dimensions[col_letter].width = column_widths[col_num-1]

    mapa_tipos = {
        "8FPH": "Phishing",
        "2CMV": "Malware",
        "8FFR": "Sitios Fraudulentos",
        "4IIA": "Ataques de Fuerza Bruta",
        "4IIV": "Ataques de Fuerza Bruta",
        "ACF": "Campaña Fraudulenta",
        "AIA": "Investigacion de Amenazas",
        "AVC": "Vulnerabilidad Critica"
    }

    contador = 1
    for alerta in alertas:
        tipo_reporte = "Otro / Desconocido"
        nombre_upper = alerta.nombre_alerta.upper()
        
        for codigo, descripcion in mapa_tipos.items():
            if codigo in nombre_upper:
                tipo_reporte = descripcion
                break
        
        if tipo_reporte == "Otro / Desconocido" and alerta.tipo_alerta:
            tipo_reporte = mapa_tipos.get(alerta.tipo_alerta, alerta.tipo_alerta)

        row_data = [
            contador,
            alerta.fecha_realizacion.strftime('%d/%m/%Y'),
            alerta.nombre_alerta,
            tipo_reporte,
            "Alta",
            "CSIRT",
            "Bloqueo de IoC",
            alerta.fecha_realizacion.strftime('%d/%m/%Y'),
            "OK"
        ]
        ws.append(row_data)
        
        current_row = ws.max_row
        for col_idx, cell in enumerate(ws[current_row], 1):
            cell.border = thin_border
            if col_idx in [3, 4, 7, 9]: 
                cell.alignment = left_align
            else:
                cell.alignment = center_align
        
        contador += 1

    excel_file = BytesIO()
    wb.save(excel_file)
    excel_file.seek(0)
    return excel_file